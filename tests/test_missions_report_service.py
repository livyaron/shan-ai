"""Pure-function tests for the חדר מבצעים XLSX report + AI focus summary.

Detached model instances only — no DB, matching test_missions_menu_service.py.
"""

import datetime
from datetime import date, timedelta
from io import BytesIO

import pytest
from sqlalchemy.orm import configure_mappers, class_mapper

from app.models import Mission, MissionUpdate, User
from app.services import missions_report_service as mrs

configure_mappers()
_mission_mgr = class_mapper(Mission).class_manager
_user_mgr = class_mapper(User).class_manager
_upd_mgr = class_mapper(MissionUpdate).class_manager

TODAY = date(2026, 7, 15)
NOW = datetime.datetime(2026, 7, 15, 6, 0)


def _make_user(**kwargs):
    defaults = dict(id=1, username="דני לוי", telegram_id=111)
    defaults.update(kwargs)
    u = _user_mgr.new_instance()
    for k, v in defaults.items():
        setattr(u, k, v)
    return u


def _make_mission(**kwargs):
    defaults = dict(
        id=1,
        title="בדיקת שנאי בתחנת שדרות",
        description=None,
        is_urgent=True,
        is_important=True,
        status="open",
        owner_id=1,
        created_by_id=1,
        due_date=None,
        created_at=datetime.datetime(2026, 7, 1, 8, 0),
        updated_at=datetime.datetime(2026, 7, 10, 8, 0),
        completed_at=None,
    )
    owner = kwargs.pop("owner", _make_user())
    updates = kwargs.pop("updates", None)
    defaults.update(kwargs)
    m = _mission_mgr.new_instance()
    for k, v in defaults.items():
        setattr(m, k, v)
    m.owner = owner
    m.created_by = owner
    # Left unset when not given, so get_mission_updates sees an unloaded log —
    # exactly what a query without selectinload produces.
    if updates is not None:
        m.updates = updates
    return m


def _make_update(**kwargs):
    """Transient MissionUpdate. created_at is naive UTC, as stored."""
    defaults = dict(
        id=1, mission_id=1, text="הוחלף מבודד", kind=None,
        author_id=1, author_name="דני",
        created_at=datetime.datetime(2026, 7, 14, 15, 6),
    )
    defaults.update(kwargs)
    u = _upd_mgr.new_instance()
    for k, v in defaults.items():
        setattr(u, k, v)
    u.author = None
    return u


# ── due_bucket boundaries ──────────────────────────────────────────────────

@pytest.mark.parametrize("delta,expected", [
    (-1, "באיחור"),
    (0, "היום"),
    (1, f"עד {mrs.AT_RISK_DAYS} ימים"),
    (3, f"עד {mrs.AT_RISK_DAYS} ימים"),
    (4, "השבוע"),
    (7, "השבוע"),
    (8, "מאוחר יותר"),
])
def test_due_bucket_boundaries(delta, expected):
    assert mrs.due_bucket(TODAY + timedelta(days=delta), TODAY) == expected


def test_due_bucket_no_date():
    assert mrs.due_bucket(None, TODAY) == "ללא תאריך יעד"


def test_bucket_order_covers_every_bucket():
    produced = {mrs.due_bucket(TODAY + timedelta(days=d), TODAY) for d in range(-2, 30)}
    produced.add(mrs.due_bucket(None, TODAY))
    assert produced == set(mrs.BUCKET_ORDER)


# ── priority weight mirrors the SQL case() ─────────────────────────────────

def test_priority_weight_ordering():
    assert mrs._priority_weight(_make_mission(is_urgent=True, is_important=True)) == 0
    assert mrs._priority_weight(_make_mission(is_urgent=False, is_important=True)) == 1
    assert mrs._priority_weight(_make_mission(is_urgent=True, is_important=False)) == 2
    assert mrs._priority_weight(_make_mission(is_urgent=False, is_important=False)) == 3


# ── stats ──────────────────────────────────────────────────────────────────

def test_stats_overdue_rate_and_buckets():
    active = [
        _make_mission(id=1, due_date=TODAY - timedelta(days=2)),
        _make_mission(id=2, due_date=TODAY),
        _make_mission(id=3, due_date=TODAY + timedelta(days=10)),
        _make_mission(id=4, due_date=None),
    ]
    stats = mrs._compute_stats(active, [], TODAY, NOW)
    assert stats["total_open"] == 4
    assert stats["overdue"] == 1
    assert stats["overdue_rate"] == 25
    assert stats["buckets"]["באיחור"] == 1
    assert stats["buckets"]["היום"] == 1
    assert stats["buckets"]["מאוחר יותר"] == 1
    assert stats["buckets"]["ללא תאריך יעד"] == 1


def test_stats_per_owner_aggregation():
    a, b = _make_user(id=1, username="דני"), _make_user(id=2, username="רותם")
    active = [
        _make_mission(id=1, owner=a, due_date=TODAY - timedelta(days=3)),
        _make_mission(id=2, owner=a, due_date=TODAY - timedelta(days=1)),
        _make_mission(id=3, owner=a, due_date=TODAY + timedelta(days=5)),
        _make_mission(id=4, owner=b, due_date=TODAY + timedelta(days=5)),
    ]
    stats = mrs._compute_stats(active, [], TODAY, NOW)
    by_name = {r["owner"]: r for r in stats["owners"]}
    assert by_name["דני"]["open"] == 3
    assert by_name["דני"]["overdue"] == 2
    assert by_name["דני"]["overdue_rate"] == 67
    assert by_name["רותם"]["overdue"] == 0
    # Most-overdue owner sorts first — that's the row a manager should see immediately.
    assert stats["owners"][0]["owner"] == "דני"


def test_stats_cycle_time_and_throughput():
    closed = [
        _make_mission(
            id=10, status="done",
            created_at=datetime.datetime(2026, 7, 1, 8, 0),
            completed_at=datetime.datetime(2026, 7, 11, 8, 0),
        ),
        _make_mission(
            id=11, status="done",
            created_at=datetime.datetime(2026, 6, 25, 8, 0),
            completed_at=datetime.datetime(2026, 6, 29, 8, 0),
        ),
    ]
    stats = mrs._compute_stats([], closed, TODAY, NOW)
    assert stats["closed_30d"] == 2
    assert stats["closed_7d"] == 1        # only the 11/07 one is inside 7 days
    assert stats["avg_cycle_30d"] == 7.0  # (10 + 4) / 2


def test_stats_stale_missions():
    active = [
        _make_mission(id=1, updated_at=datetime.datetime(2026, 5, 1, 8, 0)),
        _make_mission(id=2, updated_at=datetime.datetime(2026, 7, 14, 8, 0)),
    ]
    stats = mrs._compute_stats(active, [], TODAY, NOW)
    assert stats["stale"] == 1


def test_stats_empty_board_does_not_divide_by_zero():
    stats = mrs._compute_stats([], [], TODAY, NOW)
    assert stats["overdue_rate"] == 0
    assert stats["avg_age_open"] == 0.0
    assert stats["avg_cycle_30d"] == 0.0


# ── insights ───────────────────────────────────────────────────────────────

def test_insights_always_return_something():
    stats = mrs._compute_stats([], [], TODAY, NOW)
    out = mrs.compute_insights(stats)
    assert out and all({"sev", "icon", "headline", "action"} <= set(i) for i in out)


def test_insights_flag_concentrated_overdue_load():
    a, b = _make_user(id=1, username="דני"), _make_user(id=2, username="רותם")
    active = [_make_mission(id=i, owner=a, due_date=TODAY - timedelta(days=2))
              for i in range(1, 5)]
    active.append(_make_mission(id=9, owner=b, due_date=TODAY + timedelta(days=4)))
    stats = mrs._compute_stats(active, [], TODAY, NOW)
    assert any("דני" in i["headline"] for i in mrs.compute_insights(stats))


# ── workbook ───────────────────────────────────────────────────────────────

def _sample_data():
    a = _make_user(id=1, username="דני")
    active = [
        _make_mission(id=1, owner=a, due_date=TODAY - timedelta(days=2)),
        _make_mission(id=2, owner=a, due_date=TODAY + timedelta(days=2)),
        _make_mission(id=3, owner=a, due_date=TODAY + timedelta(days=20)),
    ]
    closed = [_make_mission(
        id=10, owner=a, status="done", due_date=TODAY - timedelta(days=5),
        created_at=datetime.datetime(2026, 7, 1, 8, 0),
        completed_at=datetime.datetime(2026, 7, 6, 8, 0),
    )]
    stats = mrs._compute_stats(active, closed, TODAY, NOW)
    return {
        "open_rows": [{
            "id": m.id, "title": m.title, "description": "", "quadrant": "בצע עכשיו",
            "urgent": "כן", "important": "כן", "status": "פתוחה", "owner": "דני",
            "created_by": "דני", "due": "01/01/2026", "days_to_due": 1,
            "days_overdue": None, "age_days": 5, "created_at": "", "updated_at": "",
            "last_status_update": "—",
            "_overdue": m.due_date < TODAY, "_at_risk": False,
        } for m in active],
        "closed_rows": [{
            "id": 10, "title": "סגורה", "quadrant": "בצע עכשיו", "status": "הושלמה",
            "owner": "דני", "due": "10/07/2026", "completed_at": "06/07/2026 08:00",
            "cycle_days": 5, "on_time": "כן", "created_at": "",
        }],
        "stats": stats,
        "meta": {"generated_at": "15/07/2026 06:00", "date_slug": "15-07-2026"},
    }


def test_build_workbook_structure():
    from openpyxl import load_workbook

    data = _sample_data()
    payload = mrs.build_workbook(data, ai_text="• תובנה לדוגמה — פעולה לדוגמה")
    assert isinstance(payload, bytes) and payload[:2] == b"PK"

    wb = load_workbook(BytesIO(payload))
    assert wb.sheetnames == [mrs.SHEET_SUMMARY, mrs.SHEET_OPEN, mrs.SHEET_CLOSED]

    # Hebrew sheets must render right-to-left or every column reads backwards.
    for name in wb.sheetnames:
        assert wb[name].sheet_view.rightToLeft is True

    ws_open = wb[mrs.SHEET_OPEN]
    assert [c.value for c in ws_open[1]] == mrs.OPEN_HEADERS
    assert ws_open.max_row == len(data["open_rows"]) + 1
    assert ws_open.freeze_panes == "A2"

    ws_closed = wb[mrs.SHEET_CLOSED]
    assert [c.value for c in ws_closed[1]] == mrs.CLOSED_HEADERS
    assert ws_closed.max_row == len(data["closed_rows"]) + 1


def test_build_workbook_survives_empty_board():
    from openpyxl import load_workbook

    stats = mrs._compute_stats([], [], TODAY, NOW)
    data = {"open_rows": [], "closed_rows": [], "stats": stats,
            "meta": {"generated_at": "15/07/2026 06:00", "date_slug": "15-07-2026"}}
    wb = load_workbook(BytesIO(mrs.build_workbook(data, ai_text="")))
    assert len(wb.sheetnames) == 3


def test_build_workbook_without_ai_text():
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(mrs.build_workbook(_sample_data(), ai_text="")))
    ws = wb[mrs.SHEET_SUMMARY]
    texts = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    # The deterministic insight block still renders when the LLM produced nothing.
    assert any("תובנות AI" in t for t in texts)
    assert any("אינן זמינות" in t for t in texts)


# ── focus summary fallback ─────────────────────────────────────────────────

def test_focus_summary_plain_fallback_lists_buckets():
    a = _make_user(id=1, username="דני")
    groups = {
        "late": [_make_mission(id=1, owner=a, due_date=TODAY - timedelta(days=2))],
        "today": [],
        "soon": [_make_mission(id=2, owner=a, due_date=TODAY + timedelta(days=2))],
        "nodate": [],
    }
    text = mrs._format_at_risk_plain(groups, TODAY)
    assert text.startswith("‏")  # RTL mark on the first line (CLAUDE.md §5)
    assert "באיחור" in text
    assert "אינו זמין" in text


def test_focus_summary_plain_fallback_when_board_clean():
    groups = {"late": [], "today": [], "soon": [], "nodate": []}
    text = mrs._format_at_risk_plain(groups, TODAY)
    assert "אין משימות באיחור" in text


async def test_build_focus_summary_falls_back_when_llm_raises(monkeypatch):
    a = _make_user(id=1, username="דני")
    groups = {
        "late": [_make_mission(id=1, owner=a, due_date=TODAY - timedelta(days=2))],
        "today": [], "soon": [], "nodate": [],
    }

    async def _fake_collect(_session):
        return groups

    async def _boom(*args, **kwargs):
        raise RuntimeError("rate limited")

    monkeypatch.setattr(mrs, "collect_at_risk", _fake_collect)
    monkeypatch.setattr(mrs.oms, "today_il", lambda: TODAY)
    from app.services import llm_router
    monkeypatch.setattr(llm_router, "llm_chat", _boom)

    text = await mrs.build_focus_summary(session=None)
    # A dead LLM must degrade to the raw list, never to an exception.
    assert "אינו זמין" in text


async def test_build_focus_summary_escapes_model_output(monkeypatch):
    a = _make_user(id=1, username="דני")
    groups = {
        "late": [_make_mission(id=1, owner=a, due_date=TODAY - timedelta(days=2))],
        "today": [], "soon": [], "nodate": [],
    }

    async def _fake_collect(_session):
        return groups

    async def _inject(*args, **kwargs):
        return "<script>alert(1)</script> הערכת מצב"

    monkeypatch.setattr(mrs, "collect_at_risk", _fake_collect)
    monkeypatch.setattr(mrs.oms, "today_il", lambda: TODAY)
    from app.services import llm_router
    monkeypatch.setattr(llm_router, "llm_chat", _inject)

    text = await mrs.build_focus_summary(session=None)
    assert "<script>" not in text
    assert "&lt;script&gt;" in text


def test_plain_fallback_has_bullets_and_todo_section():
    a, b = _make_user(id=1, username="דני"), _make_user(id=2, username="רותם")
    groups = {
        "late": [
            _make_mission(id=1, owner=a, due_date=TODAY - timedelta(days=12)),
            _make_mission(id=2, owner=a, due_date=TODAY - timedelta(days=2)),
        ],
        "today": [_make_mission(id=3, owner=b, due_date=TODAY)],
        "soon": [_make_mission(id=4, owner=b, due_date=TODAY + timedelta(days=2))],
        "nodate": [],
    }
    text = mrs._format_at_risk_plain(groups, TODAY)
    assert "🔎 תמונת מצב" in text
    assert "✅ משימות לביצוע היום" in text
    assert "🚨 לטיפול הנהלה" in text          # one mission is >7 days late
    assert text.count("•") >= 6
    # Every to-do line names its owner so the list is actionable as-is.
    assert "[דני]" in text and "[רותם]" in text


def test_decorate_summary_bolds_headers_and_normalises_bullets():
    raw = "תמונת מצב\n- שתי משימות באיחור\nמשימות לביצוע היום\n* תאם הפסקת חשמל"
    out = mrs._decorate_summary(raw)
    assert "<b>🔎 תמונת מצב</b>" in out
    assert "<b>✅ משימות לביצוע היום</b>" in out
    assert "• שתי משימות באיחור" in out
    assert "• תאם הפסקת חשמל" in out
    assert "- שתי" not in out and "* תאם" not in out


def test_decorate_summary_escapes_before_decorating():
    out = mrs._decorate_summary("תמונת מצב\n- <img src=x onerror=1>")
    assert "<img" not in out
    assert "&lt;img" in out
    assert "<b>🔎 תמונת מצב</b>" in out       # our own tags survive


# ── daily caches ───────────────────────────────────────────────────────────

@pytest.fixture(autouse=True)
def _clean_caches():
    mrs.clear_caches()
    yield
    mrs.clear_caches()


async def test_focus_summary_is_built_once_per_day(monkeypatch):
    a = _make_user(id=1, username="דני")
    groups = {"late": [_make_mission(id=1, owner=a, due_date=TODAY - timedelta(days=2))],
              "today": [], "soon": [], "nodate": []}
    calls = []

    async def _fake_collect(_session):
        return groups

    async def _llm(*args, **kwargs):
        calls.append(1)
        return "תמונת מצב\n• משימה אחת באיחור"

    monkeypatch.setattr(mrs, "collect_at_risk", _fake_collect)
    monkeypatch.setattr(mrs.oms, "today_il", lambda: TODAY)
    from app.services import llm_router
    monkeypatch.setattr(llm_router, "llm_chat", _llm)

    first = await mrs.build_focus_summary(session=None)
    second = await mrs.build_focus_summary(session=None)
    assert first == second
    assert len(calls) == 1, "second press must be served from the day cache"

    # The 04:10 prewarm is the only thing allowed to rebuild.
    await mrs.build_focus_summary(session=None, force=True)
    assert len(calls) == 2


async def test_llm_failure_is_not_cached(monkeypatch):
    """A transient Groq outage must not pin the fallback text for the rest of the day."""
    a = _make_user(id=1, username="דני")
    groups = {"late": [_make_mission(id=1, owner=a, due_date=TODAY - timedelta(days=2))],
              "today": [], "soon": [], "nodate": []}
    state = {"fail": True}

    async def _fake_collect(_session):
        return groups

    async def _llm(*args, **kwargs):
        if state["fail"]:
            raise RuntimeError("rate limited")
        return "תמונת מצב\n• הכול תקין"

    monkeypatch.setattr(mrs, "collect_at_risk", _fake_collect)
    monkeypatch.setattr(mrs.oms, "today_il", lambda: TODAY)
    from app.services import llm_router
    monkeypatch.setattr(llm_router, "llm_chat", _llm)

    assert "אינו זמין" in await mrs.build_focus_summary(session=None)
    state["fail"] = False
    assert "אינו זמין" not in await mrs.build_focus_summary(session=None)


async def test_report_bytes_cached_per_day(monkeypatch):
    builds = []

    async def _fake_collect(_session):
        builds.append(1)
        return _sample_data()

    async def _fake_ai(_data, force=False, session=None):
        return "• תובנה"

    monkeypatch.setattr(mrs, "collect_report_data", _fake_collect)
    monkeypatch.setattr(mrs, "get_ai_insights", _fake_ai)
    monkeypatch.setattr(mrs.oms, "today_il", lambda: TODAY)

    first, name1, _ = await mrs.build_report_bytes(session=None)
    second, name2, _ = await mrs.build_report_bytes(session=None)
    assert first is second and name1 == name2
    assert len(builds) == 1

    await mrs.build_report_bytes(session=None, force=True)
    assert len(builds) == 2


async def test_no_ai_download_does_not_poison_the_day_cache(monkeypatch):
    """?ai=0 is an explicit opt-out — it must not become the cached daily report."""
    async def _fake_collect(_session):
        return _sample_data()

    async def _fake_ai(_data, force=False, session=None):
        return "• תובנה מה-AI"

    monkeypatch.setattr(mrs, "collect_report_data", _fake_collect)
    monkeypatch.setattr(mrs, "get_ai_insights", _fake_ai)
    monkeypatch.setattr(mrs.oms, "today_il", lambda: TODAY)

    await mrs.build_report_bytes(session=None, with_ai=False)
    assert mrs.cache_status()["report"] is False
    await mrs.build_report_bytes(session=None)
    assert mrs.cache_status()["report"] is True


async def test_prewarm_fills_both_caches(monkeypatch):
    a = _make_user(id=1, username="דני")
    groups = {"late": [_make_mission(id=1, owner=a, due_date=TODAY - timedelta(days=2))],
              "today": [], "soon": [], "nodate": []}

    async def _fake_collect_report(_session):
        return _sample_data()

    async def _fake_collect_risk(_session):
        return groups

    async def _llm(*args, **kwargs):
        return "תמונת מצב\n• נקודה"

    monkeypatch.setattr(mrs, "collect_report_data", _fake_collect_report)
    monkeypatch.setattr(mrs, "collect_at_risk", _fake_collect_risk)
    monkeypatch.setattr(mrs.oms, "today_il", lambda: TODAY)
    from app.services import llm_router
    monkeypatch.setattr(llm_router, "llm_chat", _llm)

    assert await mrs.prewarm_daily(session=None) == {"report": True, "summary": True}
    status = mrs.cache_status()
    assert status["report"] is True and status["summary"] is True


async def test_prewarm_reports_partial_failure(monkeypatch):
    """A broken report build must still leave the summary warm, and vice versa."""
    async def _boom(_session):
        raise RuntimeError("db down")

    async def _fake_collect_risk(_session):
        return {"late": [], "today": [], "soon": [], "nodate": []}

    monkeypatch.setattr(mrs, "collect_report_data", _boom)
    monkeypatch.setattr(mrs, "collect_at_risk", _fake_collect_risk)
    monkeypatch.setattr(mrs.oms, "today_il", lambda: TODAY)

    assert await mrs.prewarm_daily(session=None) == {"report": False, "summary": True}


async def test_refresh_rebuilds_data_but_reuses_cached_ai(monkeypatch):
    """The zero-token guarantee: a refresh re-reads the board but never re-calls the LLM."""
    collects, llm_calls = [], []

    async def _fake_collect(_session):
        collects.append(1)
        return _sample_data()

    async def _llm(*args, **kwargs):
        llm_calls.append(1)
        return "\u2022 \u05ea\u05d5\u05d1\u05e0\u05d4 \u05de\u05d4-AI"

    monkeypatch.setattr(mrs, "collect_report_data", _fake_collect)
    monkeypatch.setattr(mrs.oms, "today_il", lambda: TODAY)
    from app.services import llm_router
    monkeypatch.setattr(llm_router, "llm_chat", _llm)

    await mrs.build_report_bytes(session=None)          # first build warms both caches
    assert (len(collects), len(llm_calls)) == (1, 1)

    await mrs.build_report_bytes(session=None)          # cached — no work at all
    assert (len(collects), len(llm_calls)) == (1, 1)

    await mrs.build_report_bytes(session=None, refresh_data=True)
    assert len(collects) == 2, "refresh must re-read the board"
    assert len(llm_calls) == 1, "refresh must reuse the cached AI narrative — no Groq call"


async def test_refresh_builds_ai_once_when_cache_is_cold(monkeypatch):
    """Covers a process restart before the 04:10 prewarm ran."""
    llm_calls = []

    async def _fake_collect(_session):
        return _sample_data()

    async def _llm(*args, **kwargs):
        llm_calls.append(1)
        return "\u2022 \u05ea\u05d5\u05d1\u05e0\u05d4"

    monkeypatch.setattr(mrs, "collect_report_data", _fake_collect)
    monkeypatch.setattr(mrs.oms, "today_il", lambda: TODAY)
    from app.services import llm_router
    monkeypatch.setattr(llm_router, "llm_chat", _llm)

    await mrs.build_report_bytes(session=None, refresh_data=True)
    assert len(llm_calls) == 1
    await mrs.build_report_bytes(session=None, refresh_data=True)
    assert len(llm_calls) == 1, "the narrative built by the first refresh is now cached"


async def test_refresh_replaces_the_cached_report(monkeypatch):
    """After a refresh, a plain press must serve the refreshed bytes, not the stale ones."""
    board = {"rows": 3}

    async def _fake_collect(_session):
        data = _sample_data()
        data["open_rows"] = data["open_rows"][: board["rows"]]
        return data

    async def _fake_ai(_data, force=False, session=None):
        return "\u2022 \u05ea\u05d5\u05d1\u05e0\u05d4"

    monkeypatch.setattr(mrs, "collect_report_data", _fake_collect)
    monkeypatch.setattr(mrs, "get_ai_insights", _fake_ai)
    monkeypatch.setattr(mrs.oms, "today_il", lambda: TODAY)

    stale, _, _ = await mrs.build_report_bytes(session=None)
    board["rows"] = 1                                    # somebody closed two missions
    fresh, _, _ = await mrs.build_report_bytes(session=None, refresh_data=True)
    assert fresh != stale

    served, _, _ = await mrs.build_report_bytes(session=None)
    assert served == fresh


async def test_report_returns_generated_at_stamp(monkeypatch):
    async def _fake_collect(_session):
        return _sample_data()

    async def _fake_ai(_data, force=False, session=None):
        return ""

    monkeypatch.setattr(mrs, "collect_report_data", _fake_collect)
    monkeypatch.setattr(mrs, "get_ai_insights", _fake_ai)
    monkeypatch.setattr(mrs.oms, "today_il", lambda: TODAY)

    payload, filename, generated_at = await mrs.build_report_bytes(session=None)
    assert payload[:2] == b"PK"
    assert filename.endswith(".xlsx")
    assert generated_at == "15/07/2026 06:00"


def test_cache_put_keeps_only_today():
    cache = {}
    mrs._cache_put(cache, "2026-07-14", "old")
    mrs._cache_put(cache, "2026-07-15", "new")
    assert cache == {"2026-07-15": "new"}


# ── menu keyboard guard ────────────────────────────────────────────────────

def test_menu_keyboard_exposes_report_buttons():
    from app.services.missions_menu_service import get_menu_keyboard

    kb = get_menu_keyboard({"do": 1, "plan": 0, "delegate": 0, "backlog": 0})
    tokens = [b.callback_data for row in kb.inline_keyboard for b in row]
    assert "om:xls" in tokens
    assert "om:xlsr" in tokens
    assert "om:sum" in tokens
    # Colon-free tail so the om:* dispatcher can keep using split(':').
    assert all(t.count(":") == 1 for t in ("om:xls", "om:xlsr", "om:sum"))


# ── Day cache survives a restart (the 🧠 סיכום AI latency fix) ──────────────

class _FakeCacheSession:
    """Stand-in for an AsyncSession backed by a dict of persisted rows.

    Models just enough of the mission_report_cache read/write path to prove the
    read-through works without a live Postgres.
    """

    def __init__(self, rows=None):
        self.rows = dict(rows or {})   # kind -> SimpleNamespace
        self.writes = []

    async def scalar(self, _stmt):
        return self.rows.get(self._kind_of(_stmt))

    async def execute(self, _stmt):
        return None

    async def commit(self):
        return None

    async def rollback(self):
        return None

    @staticmethod
    def _kind_of(stmt):
        # The only queries this session ever sees filter on a single kind literal.
        for kind in (mrs.KIND_SUMMARY, mrs.KIND_AI_INSIGHTS, mrs.KIND_REPORT):
            if f"'{kind}'" in str(stmt.compile(compile_kwargs={"literal_binds": True})):
                return kind
        return None


def _cached_row(**kwargs):
    from types import SimpleNamespace
    defaults = dict(text=None, blob=None, filename=None, generated_at=None)
    defaults.update(kwargs)
    return SimpleNamespace(**defaults)


async def test_summary_served_from_the_db_after_a_restart(monkeypatch):
    """The exact bug: a redeploy wiped memory and every press paid a full LLM call."""
    llm_calls = []

    async def _llm(*args, **kwargs):
        llm_calls.append(1)
        return "תמונת מצב\n• נקודה"

    monkeypatch.setattr(mrs.oms, "today_il", lambda: TODAY)
    from app.services import llm_router
    monkeypatch.setattr(llm_router, "llm_chat", _llm)

    session = _FakeCacheSession({mrs.KIND_SUMMARY: _cached_row(text="‏🧠 סיכום של הבוקר")})
    mrs.clear_caches()                       # simulate the fresh container

    text = await mrs.build_focus_summary(session)
    assert text == "‏🧠 סיכום של הבוקר"
    assert llm_calls == [], "a persisted summary must not trigger a Groq call"
    assert mrs.cache_status()["summary"] is True, "and it must land back in memory"


async def test_load_caches_from_db_restores_every_artifact(monkeypatch):
    monkeypatch.setattr(mrs.oms, "today_il", lambda: TODAY)
    session = _FakeCacheSession({
        mrs.KIND_SUMMARY: _cached_row(text="סיכום"),
        mrs.KIND_AI_INSIGHTS: _cached_row(text="תובנות"),
        mrs.KIND_REPORT: _cached_row(blob=b"xlsx", filename="a.xlsx", generated_at="15/07/2026"),
    })
    mrs.clear_caches()

    loaded = await mrs.load_caches_from_db(session)
    assert loaded == {"summary": True, "ai_insights": True, "report": True}
    status = mrs.cache_status()
    assert status["summary"] and status["ai_insights"] and status["report"]


async def test_is_summary_warm_reads_through_to_the_db(monkeypatch):
    monkeypatch.setattr(mrs.oms, "today_il", lambda: TODAY)
    mrs.clear_caches()
    assert await mrs.is_summary_warm(_FakeCacheSession()) is False
    assert await mrs.is_summary_warm(
        _FakeCacheSession({mrs.KIND_SUMMARY: _cached_row(text="סיכום")})
    ) is True


async def test_prewarm_reports_failure_when_the_llm_leg_falls_back(monkeypatch):
    """The silent bug: the prewarm logged summary=True while the cache stayed empty.

    A prewarm that fell back must report False so the watchdog retries, instead
    of leaving the button cold — and slow — for the rest of the day.
    """
    groups = {"late": [_make_mission(id=1, due_date=TODAY - timedelta(days=2))],
              "today": [], "soon": [], "nodate": []}

    async def _fake_collect_report(_session):
        return _sample_data()

    async def _fake_collect_risk(_session):
        return groups

    async def _llm_down(*args, **kwargs):
        raise RuntimeError("groq rate limited")

    monkeypatch.setattr(mrs, "collect_report_data", _fake_collect_report)
    monkeypatch.setattr(mrs, "collect_at_risk", _fake_collect_risk)
    monkeypatch.setattr(mrs.oms, "today_il", lambda: TODAY)
    from app.services import llm_router
    monkeypatch.setattr(llm_router, "llm_chat", _llm_down)

    result = await mrs.prewarm_daily(session=None)
    assert result["summary"] is False
    assert mrs.cache_status()["summary"] is False


async def test_focus_summary_still_returns_the_plain_list_for_normal_callers(monkeypatch):
    """strict=False (every user-facing caller) must never raise — text or nothing."""
    groups = {"late": [_make_mission(id=1, due_date=TODAY - timedelta(days=2))],
              "today": [], "soon": [], "nodate": []}

    async def _fake_collect_risk(_session):
        return groups

    async def _llm_down(*args, **kwargs):
        raise RuntimeError("groq down")

    monkeypatch.setattr(mrs, "collect_at_risk", _fake_collect_risk)
    monkeypatch.setattr(mrs.oms, "today_il", lambda: TODAY)
    from app.services import llm_router
    monkeypatch.setattr(llm_router, "llm_chat", _llm_down)

    text = await mrs.build_focus_summary(session=None)
    assert "סיכום ה-AI אינו זמין" in text
    assert mrs.cache_status()["summary"] is False, "a transient outage must not pin the fallback"


async def test_at_risk_prompt_is_capped(monkeypatch):
    """An unbounded prompt is what made a cold build slow on a busy board."""
    monkeypatch.setattr(mrs.oms, "today_il", lambda: TODAY)
    late = [_make_mission(id=i, title=f"משימה {i:03d}", due_date=TODAY - timedelta(days=3))
            for i in range(1, mrs.PROMPT_MISSION_CAP + 21)]
    context = mrs._at_risk_context(
        {"late": late, "today": [], "soon": [], "nodate": []}, TODAY
    )
    assert context.count("| אחראי:") == mrs.PROMPT_MISSION_CAP
    # The owner-load header is still computed from the whole board.
    assert f"{len(late)} משימות" in context


# ── "עדכון אחרון" — when each owner last reported on their open work ────────

def _owner_stats(active, closed=()):
    return mrs._compute_stats(list(active), list(closed), TODAY, NOW)


def test_last_update_at_reads_the_status_log_not_updated_at():
    """Editing a due date bumps updated_at; it is not a report on the work."""
    m = _make_mission(updated_at=datetime.datetime(2026, 7, 15, 9, 0), updates=[])
    assert mrs._last_update_at(m) is None

    m = _make_mission(updates=[
        _make_update(id=1, created_at=datetime.datetime(2026, 7, 10, 8, 0)),
        _make_update(id=2, created_at=datetime.datetime(2026, 7, 12, 8, 0)),
    ])
    assert mrs._last_update_at(m) == datetime.datetime(2026, 7, 12, 8, 0)


def test_last_update_at_returns_none_when_the_log_is_not_loaded():
    """A lazy load here would raise MissingGreenlet under asyncio."""
    assert mrs._last_update_at(_make_mission()) is None


def test_owner_row_carries_the_newest_update_across_their_open_missions():
    a = _make_user(id=1, username="דני")
    stats = _owner_stats([
        _make_mission(id=1, owner=a, updates=[
            _make_update(id=1, created_at=datetime.datetime(2026, 7, 9, 8, 0))]),
        _make_mission(id=2, owner=a, updates=[
            _make_update(id=2, created_at=datetime.datetime(2026, 7, 13, 8, 0))]),
        _make_mission(id=3, owner=a, updates=[]),
    ])
    row = stats["owners"][0]
    assert row["last_update"] == "13/07/2026"
    assert row["days_since_update"] == 2          # TODAY is 15/07/2026


def test_owner_with_no_reports_shows_a_dash():
    a = _make_user(id=1, username="דני")
    row = _owner_stats([_make_mission(id=1, owner=a, updates=[])])["owners"][0]
    assert row["last_update"] == "—"
    assert row["days_since_update"] is None


def test_a_report_on_a_closed_mission_does_not_count_as_reporting():
    """The column answers "is this person reporting on their OPEN work?"."""
    a = _make_user(id=1, username="דני")
    closed = _make_mission(
        id=10, owner=a, status="done",
        completed_at=datetime.datetime(2026, 7, 14, 8, 0),
        updates=[_make_update(id=9, created_at=datetime.datetime(2026, 7, 14, 8, 0))],
    )
    stats = _owner_stats([_make_mission(id=1, owner=a, updates=[])], [closed])
    row = stats["owners"][0]
    assert row["last_update"] == "—"
    assert row["done_30"] == 1


def test_silent_owners_counts_only_those_holding_open_work():
    fresh = _make_user(id=1, username="דני")
    quiet = _make_user(id=2, username="רות")
    stats = _owner_stats([
        _make_mission(id=1, owner=fresh, updates=[
            _make_update(id=1, created_at=datetime.datetime(2026, 7, 14, 8, 0))]),
        _make_mission(id=2, owner=quiet, updates=[
            _make_update(id=2, created_at=datetime.datetime(2026, 5, 1, 8, 0))]),
    ])
    assert stats["silent_owners"] == 1


def test_owner_closed_out_is_not_silent():
    """Someone with nothing open cannot be 'not reporting'."""
    a = _make_user(id=1, username="דני")
    stats = _owner_stats([], [_make_mission(
        id=10, owner=a, status="done",
        completed_at=datetime.datetime(2026, 7, 14, 8, 0), updates=[])])
    assert stats["silent_owners"] == 0


def test_silent_owners_surfaces_as_an_insight():
    a = _make_user(id=1, username="דני")
    stats = _owner_stats([_make_mission(id=1, owner=a, updates=[])])
    headlines = " ".join(i["headline"] for i in mrs.compute_insights(stats))
    assert "לא דיווחו" in headlines


def test_stats_block_tells_the_model_how_long_each_owner_has_been_quiet():
    a = _make_user(id=1, username="דני")
    stats = _owner_stats([_make_mission(id=1, owner=a, updates=[
        _make_update(id=1, created_at=datetime.datetime(2026, 7, 13, 8, 0))])])
    block = mrs._stats_block(stats)
    assert "עדכון אחרון" in block
    assert "לפני 2 ימים" in block


def test_workbook_owner_table_has_the_last_update_column():
    from openpyxl import load_workbook

    a = _make_user(id=1, username="דני")
    active = [_make_mission(id=1, owner=a, due_date=TODAY - timedelta(days=2), updates=[
        _make_update(id=1, created_at=datetime.datetime(2026, 7, 13, 8, 0))])]
    data = {
        "open_rows": [], "closed_rows": [],
        "stats": _owner_stats(active),
        "meta": {"generated_at": "15/07/2026 06:00", "date_slug": "15-07-2026"},
    }
    wb = load_workbook(BytesIO(mrs.build_workbook(data)))
    ws = wb[mrs.SHEET_SUMMARY]
    rows = [[c.value for c in r] for r in ws.iter_rows(max_col=8)]

    hdr = next(i for i, r in enumerate(rows) if r[0] == "אחראי")
    assert rows[hdr][:7] == [
        "אחראי", "פתוחות", "באיחור", "% איחור", "הושלמו 30 יום",
        "ממוצע ימי ביצוע", "עדכון אחרון",
    ]
    assert rows[hdr][7] is None, "the table must not spill into an 8th column"
    assert rows[hdr + 1][0] == "דני"
    assert rows[hdr + 1][6] == "13/07/2026"

    flat = [c for r in rows for c in r if isinstance(c, str)]
    assert any("אחראים ללא דיווח" in c for c in flat), "the KPI block gained the figure"


# ── Per-mission "עדכון סטטוס אחרון" on the open sheet ──────────────────────

async def _collected(missions, monkeypatch):
    """Run the real collect_report_data over a fixed mission list."""
    class _FakeSession:
        async def scalars(self, _stmt):
            return _Scalars(missions)

    class _Scalars:
        def __init__(self, rows):
            self._rows = rows

        def all(self):
            return self._rows

    monkeypatch.setattr(mrs.oms, "today_il", lambda: TODAY)
    return await mrs.collect_report_data(_FakeSession())


async def test_open_row_carries_the_last_status_update(monkeypatch):
    data = await _collected([_make_mission(id=1, updates=[
        _make_update(id=1, created_at=datetime.datetime(2026, 7, 13, 6, 30)),
        _make_update(id=2, created_at=datetime.datetime(2026, 7, 14, 12, 5)),
    ])], monkeypatch)
    # 12:05 UTC → 15:05 Israel local.
    assert data["open_rows"][0]["last_status_update"] == "14/07/2026 15:05"


async def test_open_row_shows_a_dash_when_nobody_reported(monkeypatch):
    data = await _collected([_make_mission(id=1, updates=[])], monkeypatch)
    assert data["open_rows"][0]["last_status_update"] == "—"


async def test_last_status_update_is_independent_of_updated_at(monkeypatch):
    """A due-date edit bumps updated_at but must not look like a status report."""
    data = await _collected([_make_mission(
        id=1, updated_at=datetime.datetime(2026, 7, 15, 5, 0), updates=[])], monkeypatch)
    row = data["open_rows"][0]
    assert row["updated_at"] == "15/07/2026 08:00"     # 05:00 UTC → 08:00 IL
    assert row["last_status_update"] == "—"


def test_open_sheet_appends_the_column_without_shifting_the_others():
    """Saved filters and column references depend on the existing positions."""
    assert mrs.OPEN_HEADERS[-1] == "עדכון סטטוס אחרון"
    assert mrs.OPEN_HEADERS[-2] == "עודכנה לאחרונה"
    assert len(mrs.OPEN_HEADERS) == len(mrs.OPEN_WIDTHS)


def test_workbook_open_sheet_writes_the_new_column():
    from openpyxl import load_workbook

    data = _sample_data()
    data["open_rows"][0]["last_status_update"] = "14/07/2026 15:05"
    ws = load_workbook(BytesIO(mrs.build_workbook(data)))[mrs.SHEET_OPEN]

    assert [c.value for c in ws[1]] == mrs.OPEN_HEADERS
    col = mrs.OPEN_HEADERS.index("עדכון סטטוס אחרון") + 1
    assert ws.cell(2, col).value == "14/07/2026 15:05"
    # The filter must span the new column too, or it is invisible to filtering.
    assert ws.auto_filter.ref.split(":")[1].startswith(
        __import__("openpyxl").utils.get_column_letter(len(mrs.OPEN_HEADERS))
    )


# ── Timestamps are Israel-local, matching the sheet's own header ───────────

def test_fmt_dt_converts_stored_utc_to_israel_local():
    assert mrs._fmt_dt(datetime.datetime(2026, 7, 14, 12, 5)) == "14/07/2026 15:05"


def test_fmt_dt_handles_a_date_rollover():
    """22:30 UTC is already the next day in Israel."""
    assert mrs._fmt_dt(datetime.datetime(2026, 7, 14, 22, 30)) == "15/07/2026 01:30"


def test_fmt_dt_none_is_blank():
    assert mrs._fmt_dt(None) == ""
