"""Operations Room (חדר מבצעים) reporting — XLSX board report + AI focus summary.

Pure service layer, same contract as missions_menu_service: never imports the bot,
callers do the sends. Consumed by the Telegram menu (om:xls / om:sum), the web
war-room router, and the daily cron.

Nothing here touches the schema — every figure is derived from existing `missions`
columns, so there is no ALTER TABLE to run on Railway.
"""

import datetime
import html as _html
import io
import logging

from sqlalchemy import select, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models import Mission, MissionStatusEnum, User, RoleEnum
from app.services import missions_menu_service as oms

logger = logging.getLogger(__name__)

# A mission is "כמעט באיחור" when its due date is within this many days.
AT_RISK_DAYS = 3

# Active missions untouched for this long are flagged as stale on the summary sheet.
STALE_DAYS = 30

# Ceiling on how many missions are spelled out in the AI summary prompt. The
# owner-load header above the list is computed from the whole board either way,
# so the model still sees the shape of the backlog.
PROMPT_MISSION_CAP = 40

CLOSED_STATUSES = [MissionStatusEnum.DONE.value, MissionStatusEnum.CANCELLED.value]

# War-room operators: admins + the manager roles that already receive board-wide
# totals in the morning digest (eval_cron._missions_daily_digest).
MANAGER_ROLES = (
    RoleEnum.DEPARTMENT_MANAGER,
    RoleEnum.DEPUTY_DIVISION_MANAGER,
    RoleEnum.DIVISION_MANAGER,
)

OPEN_HEADERS = [
    "מזהה", "כותרת", "תיאור", "רביע", "דחוף", "חשוב", "סטטוס", "אחראי", "נוצר ע\"י",
    "תאריך יעד", "ימים ליעד", "ימים באיחור", "גיל המשימה (ימים)", "נוצרה בתאריך",
    "עודכנה לאחרונה", "עדכון סטטוס אחרון",
]
# New columns go on the END: inserting one would silently shift every saved
# filter and column reference someone has built on top of this sheet.
OPEN_WIDTHS = [8, 42, 50, 20, 8, 8, 14, 18, 18, 14, 12, 14, 18, 16, 16, 18]

CLOSED_HEADERS = [
    "מזהה", "כותרת", "רביע", "סטטוס", "אחראי", "תאריך יעד", "הושלמה בתאריך",
    "ימי ביצוע", "עמדה ביעד?", "נוצרה בתאריך",
]
CLOSED_WIDTHS = [8, 42, 20, 14, 18, 14, 16, 12, 14, 16]

SHEET_SUMMARY = "סיכום ותובנות"
SHEET_OPEN = "משימות פתוחות"
SHEET_CLOSED = "משימות סגורות"

# ── Daily caches ───────────────────────────────────────────────────────────
# Everything expensive is built once per calendar day by the 04:10 prewarm cron
# and served for the rest of the day, so a button press is a lookup rather than
# a board scan + Groq round-trip. Each cache holds only the current day's entry.
#
# Memory alone was not enough: the container is recycled on every Railway deploy,
# and the 04:10 cron does not fire again until tomorrow (and job_guard would
# refuse it if it did) — so a mid-day restart left the 🧠 סיכום AI button cold,
# and paying the full cost, for the rest of the day. Every entry is therefore
# mirrored into the mission_report_cache table and read back on a memory miss.
_ai_cache: dict[str, str] = {}                       # date -> AI insights narrative
_report_cache: dict[str, tuple[bytes, str, str]] = {} # date -> (xlsx bytes, filename, stamp)
_summary_cache: dict[str, str] = {}                  # date -> focus summary HTML

KIND_AI_INSIGHTS = "ai_insights"
KIND_REPORT = "report"
KIND_SUMMARY = "summary"


def _cache_put(cache: dict, key: str, value):
    """Store today's entry and drop yesterday's — these caches never grow."""
    cache.clear()
    cache[key] = value
    return value


async def _db_cache_get(session: AsyncSession, kind: str):
    """Today's persisted artifact, or None. Never raises — a cache is not the truth."""
    if session is None:
        return None
    from app.models import MissionReportCache
    try:
        return await session.scalar(
            select(MissionReportCache).where(
                MissionReportCache.day == oms.today_il(),
                MissionReportCache.kind == kind,
            )
        )
    except Exception as e:
        logger.warning(f"missions_report: cache read failed for {kind}: {e}")
        try:
            await session.rollback()
        except Exception:
            pass
        return None


async def _db_cache_put(
    session: AsyncSession, kind: str,
    text: str | None = None, blob: bytes | None = None,
    filename: str | None = None, generated_at: str | None = None,
) -> None:
    """Upsert today's artifact and drop older days. Never raises."""
    if session is None:
        return
    from sqlalchemy import delete as _delete
    from sqlalchemy.dialects.postgresql import insert as pg_insert
    from app.models import MissionReportCache
    day = oms.today_il()
    values = dict(day=day, kind=kind, text=text, blob=blob,
                  filename=filename, generated_at=generated_at)
    try:
        await session.execute(
            pg_insert(MissionReportCache)
            .values(**values)
            .on_conflict_do_update(
                index_elements=["day", "kind"],
                set_={k: v for k, v in values.items() if k not in ("day", "kind")},
            )
        )
        # Yesterday's rows are dead weight — this table only ever serves today.
        await session.execute(
            _delete(MissionReportCache).where(MissionReportCache.day < day)
        )
        await session.commit()
    except Exception as e:
        logger.warning(f"missions_report: cache write failed for {kind}: {e}")
        try:
            await session.rollback()
        except Exception:
            pass


def cache_status() -> dict:
    """Which of today's artifacts are already warm in memory (diagnostics)."""
    key = oms.today_il().isoformat()
    return {
        "date": key,
        "ai_insights": key in _ai_cache,
        "report": key in _report_cache,
        "summary": key in _summary_cache,
    }


async def load_caches_from_db(session: AsyncSession) -> dict:
    """Pull today's persisted artifacts into memory. Called at startup.

    This is what makes a redeploy cheap: the button stays instant without
    re-running the board scan or the Groq call.
    """
    key = oms.today_il().isoformat()
    loaded = {KIND_SUMMARY: False, KIND_AI_INSIGHTS: False, KIND_REPORT: False}

    row = await _db_cache_get(session, KIND_SUMMARY)
    if row is not None and row.text:
        _cache_put(_summary_cache, key, row.text)
        loaded[KIND_SUMMARY] = True

    row = await _db_cache_get(session, KIND_AI_INSIGHTS)
    if row is not None and row.text:
        _cache_put(_ai_cache, key, row.text)
        loaded[KIND_AI_INSIGHTS] = True

    row = await _db_cache_get(session, KIND_REPORT)
    if row is not None and row.blob:
        _cache_put(_report_cache, key,
                   (row.blob, row.filename or "", row.generated_at or ""))
        loaded[KIND_REPORT] = True

    return loaded


async def is_summary_warm(session: AsyncSession) -> bool:
    """True when today's AI summary is available without an LLM call."""
    if oms.today_il().isoformat() in _summary_cache:
        return True
    row = await _db_cache_get(session, KIND_SUMMARY)
    return bool(row is not None and row.text)


def clear_caches() -> None:
    for cache in (_ai_cache, _report_cache, _summary_cache):
        cache.clear()


# ── Helpers ────────────────────────────────────────────────────────────────

def _priority_weight(m: Mission) -> int:
    """Python mirror of missions_menu_service._priority_order()."""
    if m.is_urgent and m.is_important:
        return 0
    if m.is_important:
        return 1
    if m.is_urgent:
        return 2
    return 3


def _user_name(u: User | None) -> str:
    if u is None:
        return "—"
    return u.username or f"#{u.id}"


def _days_until(due: datetime.date | None, today: datetime.date) -> int | None:
    return None if due is None else (due - today).days


def _fmt_dt(dt: datetime.datetime | None) -> str:
    """Naive-UTC timestamp (as stored) → Israel-local 'DD/MM/YYYY HH:MM'.

    Digits are direction-neutral, so this survives RTL. The conversion matters:
    rendering the stored UTC raw made every timestamp in the sheet disagree with
    the sheet's own "הופק בתאריך" header (already Israel-local) and with the same
    event on the Telegram card by the UTC offset.
    """
    if dt is None:
        return ""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=datetime.timezone.utc)
    return dt.astimezone(oms._IL_TZ).strftime("%d/%m/%Y %H:%M")


def _age_days(created: datetime.datetime | None, today: datetime.date) -> int | None:
    return None if created is None else (today - created.date()).days


def _last_update_at(m: Mission) -> datetime.datetime | None:
    """When someone last reported on this mission, or None if nobody ever has.

    This is the status-update log, not `updated_at` — editing a due date is not
    a report on the work.
    """
    stamps = [u.created_at for u in oms.get_mission_updates(m) if u.created_at]
    return max(stamps) if stamps else None


def due_bucket(due: datetime.date | None, today: datetime.date) -> str:
    """Which urgency bucket a due date falls into. Boundaries are inclusive at 0 and AT_RISK_DAYS."""
    if due is None:
        return "ללא תאריך יעד"
    delta = (due - today).days
    if delta < 0:
        return "באיחור"
    if delta == 0:
        return "היום"
    if delta <= AT_RISK_DAYS:
        return f"עד {AT_RISK_DAYS} ימים"
    if delta <= 7:
        return "השבוע"
    return "מאוחר יותר"


BUCKET_ORDER = ["באיחור", "היום", f"עד {AT_RISK_DAYS} ימים", "השבוע", "מאוחר יותר", "ללא תאריך יעד"]


def _sanitize(text: str | None) -> str:
    """Gershayim swap before any LLM call — straight quotes break the JSON reply (CLAUDE.md §5)."""
    if not text:
        return ""
    return text.replace('"', "״").replace("'", "׳")


# ── Recipients ─────────────────────────────────────────────────────────────

async def list_war_room_operators(session: AsyncSession) -> list[User]:
    """The people who run the war room: admins + manager roles, reachable on Telegram."""
    rows = await session.scalars(
        select(User)
        .where(
            User.telegram_id.isnot(None),
            or_(User.is_admin.is_(True), User.role.in_(MANAGER_ROLES)),
        )
        .order_by(User.id)
    )
    return list(rows.all())


# ── Data collection ────────────────────────────────────────────────────────

def _description_with_updates(m: Mission) -> str:
    """Report description cell — the mission description plus its status-update log."""
    parts = [p for p in (m.description or "", oms.format_updates_block(m, html=False)) if p]
    return "\n".join(parts)


async def collect_report_data(session: AsyncSession) -> dict:
    """One pass over the whole board → flattened rows + every figure the summary sheet shows."""
    today = oms.today_il()
    now = datetime.datetime.utcnow()

    missions = list((await session.scalars(
        select(Mission).options(
            selectinload(Mission.owner),
            selectinload(Mission.created_by),
            selectinload(Mission.updates),
        )
    )).all())

    active = [m for m in missions if m.status in oms.ACTIVE_STATUSES]
    closed = [m for m in missions if m.status in CLOSED_STATUSES]

    # Overdue first, then Eisenhower weight, then soonest due date (undated last).
    active.sort(key=lambda m: (
        0 if oms.is_overdue(m, today) else 1,
        _priority_weight(m),
        m.due_date or datetime.date.max,
        -(m.id or 0),
    ))
    closed.sort(key=lambda m: (m.completed_at or datetime.datetime.min), reverse=True)

    open_rows = []
    for m in active:
        delta = _days_until(m.due_date, today)
        open_rows.append({
            "id": m.id,
            "title": m.title or "",
            "description": _description_with_updates(m),
            "quadrant": oms.quadrant_label(oms.quadrant_key(m)),
            "urgent": "כן" if m.is_urgent else "לא",
            "important": "כן" if m.is_important else "לא",
            "status": oms.STATUS_LABELS.get(m.status, m.status),
            "owner": _user_name(m.owner),
            "created_by": _user_name(m.created_by),
            "due": oms.format_due(m.due_date) if m.due_date else "",
            "days_to_due": delta,
            "days_overdue": -delta if (delta is not None and delta < 0) else None,
            "age_days": _age_days(m.created_at, today),
            "created_at": _fmt_dt(m.created_at),
            "updated_at": _fmt_dt(m.updated_at),
            # Distinct from updated_at on purpose: that moves on any edit, this
            # only when someone actually reported on the work. "—" means nobody has.
            "last_status_update": _fmt_dt(_last_update_at(m)) or "—",
            "_overdue": oms.is_overdue(m, today),
            "_at_risk": delta is not None and 0 <= delta <= AT_RISK_DAYS,
        })

    closed_rows = []
    for m in closed:
        cycle = None
        if m.completed_at and m.created_at:
            cycle = max((m.completed_at - m.created_at).days, 0)
        on_time = "ללא יעד"
        if m.due_date and m.completed_at:
            on_time = "כן" if m.completed_at.date() <= m.due_date else "לא"
        closed_rows.append({
            "id": m.id,
            "title": m.title or "",
            "quadrant": oms.quadrant_label(oms.quadrant_key(m)),
            "status": oms.STATUS_LABELS.get(m.status, m.status),
            "owner": _user_name(m.owner),
            "due": oms.format_due(m.due_date) if m.due_date else "",
            "completed_at": _fmt_dt(m.completed_at),
            "cycle_days": cycle,
            "on_time": on_time,
            "created_at": _fmt_dt(m.created_at),
        })

    stats = _compute_stats(active, closed, today, now)
    return {
        "open_rows": open_rows,
        "closed_rows": closed_rows,
        "stats": stats,
        "meta": {
            "generated_at": datetime.datetime.now(oms._IL_TZ).strftime("%d/%m/%Y %H:%M"),
            "date_slug": today.strftime("%d-%m-%Y"),
        },
    }


def _compute_stats(
    active: list[Mission],
    closed: list[Mission],
    today: datetime.date,
    now: datetime.datetime,
) -> dict:
    total_open = len(active)
    overdue = [m for m in active if oms.is_overdue(m, today)]

    buckets = {b: 0 for b in BUCKET_ORDER}
    for m in active:
        buckets[due_bucket(m.due_date, today)] += 1

    quadrants = {key: 0 for key, *_ in oms.QUADRANTS}
    for m in active:
        quadrants[oms.quadrant_key(m)] += 1

    ages = [a for a in (_age_days(m.created_at, today) for m in active) if a is not None]
    avg_age = round(sum(ages) / len(ages), 1) if ages else 0.0

    cutoff_30 = now - datetime.timedelta(days=30)
    cutoff_7 = now - datetime.timedelta(days=7)
    done_30 = [m for m in closed
               if m.status == MissionStatusEnum.DONE.value
               and m.completed_at and m.completed_at >= cutoff_30]
    done_7 = [m for m in done_30 if m.completed_at and m.completed_at >= cutoff_7]
    cycles = [max((m.completed_at - m.created_at).days, 0)
              for m in done_30 if m.completed_at and m.created_at]
    avg_cycle = round(sum(cycles) / len(cycles), 1) if cycles else 0.0

    created_7 = len([m for m in (active + closed) if m.created_at and m.created_at >= cutoff_7])
    stale_cutoff = now - datetime.timedelta(days=STALE_DAYS)
    stale = [m for m in active if m.updated_at and m.updated_at < stale_cutoff]

    # Per-owner load — the row that exposes who is actually holding the overdue work.
    per_owner: dict[str, dict] = {}

    def _owner_row(m: Mission) -> dict:
        return per_owner.setdefault(_user_name(m.owner), {
            "owner": _user_name(m.owner), "open": 0,
            "overdue": 0, "done_30": 0, "cycles": [], "last_update": None,
        })

    for m in active:
        row = _owner_row(m)
        row["open"] += 1
        if oms.is_overdue(m, today):
            row["overdue"] += 1
        # Deliberately across the owner's ACTIVE missions only: a report filed on
        # something they closed last month says nothing about their open work.
        last = _last_update_at(m)
        if last and (row["last_update"] is None or last > row["last_update"]):
            row["last_update"] = last
    for m in done_30:
        row = _owner_row(m)
        row["done_30"] += 1
        if m.completed_at and m.created_at:
            row["cycles"].append(max((m.completed_at - m.created_at).days, 0))

    owners = []
    silent_owners = 0
    for row in per_owner.values():
        cyc = row.pop("cycles")
        row["overdue_rate"] = round(row["overdue"] / row["open"] * 100) if row["open"] else 0
        row["avg_cycle"] = round(sum(cyc) / len(cyc), 1) if cyc else 0.0
        last = row.pop("last_update")
        row["last_update"] = oms.format_due(last.date()) if last else "—"
        row["days_since_update"] = (today - last.date()).days if last else None
        if row["open"] and (row["days_since_update"] is None
                            or row["days_since_update"] > STALE_DAYS):
            silent_owners += 1
        owners.append(row)
    owners.sort(key=lambda r: (-r["overdue"], -r["open"], r["owner"]))

    return {
        "total_open": total_open,
        "overdue": len(overdue),
        "overdue_rate": round(len(overdue) / total_open * 100) if total_open else 0,
        "buckets": buckets,
        "quadrants": quadrants,
        "avg_age_open": avg_age,
        "avg_cycle_30d": avg_cycle,
        "closed_7d": len(done_7),
        "closed_30d": len(done_30),
        "created_7d": created_7,
        "stale": len(stale),
        "silent_owners": silent_owners,
        "owners": owners,
    }


# ── Deterministic insights (always render, no LLM) ─────────────────────────

def compute_insights(stats: dict) -> list[dict]:
    """Rule-based insights, same shape as project_report_service._compute_insights."""
    out: list[dict] = []
    total = stats["total_open"]

    if stats["overdue"]:
        sev = "high" if stats["overdue_rate"] >= 30 else "medium"
        out.append({
            "sev": sev, "icon": "🔴",
            "headline": f"{stats['overdue']} משימות באיחור ({stats['overdue_rate']}% מהפתוחות)",
            "action": "עבור על גיליון המשימות הפתוחות — השורות האדומות בראש הרשימה.",
        })

    top = stats["owners"][0] if stats["owners"] else None
    if top and top["overdue"] >= 3 and stats["overdue"] and \
            top["overdue"] / stats["overdue"] >= 0.5:
        out.append({
            "sev": "high", "icon": "⚠️",
            "headline": f"{top['owner']} מחזיק {top['overdue']} מתוך {stats['overdue']} המשימות באיחור",
            "action": "שקול האצלה מחדש או הארכת יעדים — עומס מרוכז אצל אדם אחד.",
        })

    if stats["created_7d"] > stats["closed_7d"]:
        out.append({
            "sev": "medium", "icon": "📈",
            "headline": f"נפתחו {stats['created_7d']} משימות השבוע מול {stats['closed_7d']} שנסגרו",
            "action": "קצב הפתיחה עולה על קצב הסגירה — הלוח גדל.",
        })
    elif stats["closed_7d"] > stats["created_7d"]:
        out.append({
            "sev": "low", "icon": "✅",
            "headline": f"נסגרו {stats['closed_7d']} משימות השבוע מול {stats['created_7d']} שנפתחו",
            "action": "קצב הסגירה גבוה מקצב הפתיחה — הלוח מצטמצם.",
        })

    if stats["stale"]:
        out.append({
            "sev": "medium", "icon": "🕸",
            "headline": f"{stats['stale']} משימות פעילות ללא עדכון מעל {STALE_DAYS} ימים",
            "action": "בדוק אם הן עדיין רלוונטיות — או סגור אותן.",
        })

    if stats.get("silent_owners"):
        out.append({
            "sev": "medium", "icon": "🔇",
            "headline": (f"{stats['silent_owners']} אחראים עם משימות פתוחות "
                         f"לא דיווחו מעל {STALE_DAYS} ימים"),
            "action": "ראה עמודת «עדכון אחרון» בטבלת העומס — בקש מהם עדכון סטטוס.",
        })

    do_now = stats["quadrants"].get("do", 0)
    if do_now:
        out.append({
            "sev": "high" if do_now >= 5 else "medium", "icon": "🔥",
            "headline": f"{do_now} משימות ברביע «בצע עכשיו» (דחוף וחשוב)",
            "action": "אלה המשימות לפתוח איתן את היום.",
        })

    backlog = stats["quadrants"].get("backlog", 0)
    if total and backlog / total >= 0.4:
        out.append({
            "sev": "low", "icon": "🧺",
            "headline": f"{backlog} משימות ({round(backlog / total * 100)}%) יושבות במאגר המשימות",
            "action": "מאגר תפוח מסתיר עדיפויות — נקה או תעדף מחדש.",
        })

    if stats["avg_cycle_30d"]:
        out.append({
            "sev": "low", "icon": "⏱",
            "headline": f"זמן ביצוע ממוצע ב-30 הימים האחרונים: {stats['avg_cycle_30d']} ימים",
            "action": "השווה לדוח הקודם כדי לראות אם הקצב משתפר.",
        })

    if not out:
        out.append({
            "sev": "low", "icon": "🟢",
            "headline": "אין ממצאים חריגים — הלוח נקי",
            "action": "אין משימות באיחור ואין עומס מרוכז.",
        })
    return out


# ── AI insights (day-cached) ───────────────────────────────────────────────

_INSIGHTS_PROMPT = """אתה אנליסט תפעולי של חדר מבצעים באגף תשתיות חשמל.
לפניך נתוני לוח משימות (מודל אייזנהאואר: דחוף × חשוב).

תפקידך: לכתוב 4-6 תובנות ניהוליות קצרות בעברית, כל אחת בשורה נפרדת בפורמט:
• <התובנה> — <הפעולה המומלצת>

כללים:
- התבסס אך ורק על המספרים שניתנו. אל תמציא נתונים.
- כל תובנה חייבת להצביע על משהו שדורש החלטה, לא לחזור על המספר.
- כתוב בעברית בלבד, ללא markdown וללא כותרות.
- חשוב ביותר: אל תשתמש במרכאות כפולות (") בתוך הטקסט — השתמש בגרש בודד (׳) בלבד.
"""


def _stats_block(stats: dict) -> str:
    """Compact Hebrew rendering of the figures — sent instead of raw rows to keep tokens small."""
    lines = [
        f"סה\"כ משימות פתוחות: {stats['total_open']}",
        f"באיחור: {stats['overdue']} ({stats['overdue_rate']}%)",
        f"גיל ממוצע של משימה פתוחה: {stats['avg_age_open']} ימים",
        f"זמן ביצוע ממוצע (30 יום): {stats['avg_cycle_30d']} ימים",
        f"נסגרו ב-7 ימים: {stats['closed_7d']} | נסגרו ב-30 יום: {stats['closed_30d']}",
        f"נפתחו ב-7 ימים: {stats['created_7d']}",
        f"אחראים עם משימות פתוחות שלא דיווחו מעל {STALE_DAYS} ימים: {stats['silent_owners']}",
        f"ללא עדכון מעל {STALE_DAYS} ימים: {stats['stale']}",
        "",
        "פילוח לפי רביע:",
    ]
    for key, emoji, verb, axis in oms.QUADRANTS:
        lines.append(f"- {verb} ({axis}): {stats['quadrants'].get(key, 0)}")
    lines.append("")
    lines.append("פילוח לפי תאריך יעד:")
    for b in BUCKET_ORDER:
        lines.append(f"- {b}: {stats['buckets'].get(b, 0)}")
    lines.append("")
    lines.append("עומס לפי אחראי (פתוחות / באיחור / אחוז איחור / הושלמו ב-30 יום / עדכון אחרון):")
    for row in stats["owners"][:12]:
        days = row.get("days_since_update")
        when = "לא דיווח מעולם" if days is None else (
            "היום" if days == 0 else f"לפני {days} ימים"
        )
        lines.append(
            f"- {_sanitize(row['owner'])}: {row['open']} / {row['overdue']} / "
            f"{row['overdue_rate']}% / {row['done_30']} / {when}"
        )
    return "\n".join(lines)


async def get_ai_insights(
    data: dict, force: bool = False, session: AsyncSession | None = None,
) -> str:
    """Hebrew narrative for the summary sheet. One LLM call per calendar day, cached.

    Returns "" on any failure — the workbook must never fail because the LLM did.
    """
    key = oms.today_il().isoformat()
    if not force:
        if key in _ai_cache:
            return _ai_cache[key]
        row = await _db_cache_get(session, KIND_AI_INSIGHTS)
        if row is not None and row.text:
            return _cache_put(_ai_cache, key, row.text)

    from app.services.llm_router import llm_chat
    try:
        text = await llm_chat(
            "missions_insights",
            [
                {"role": "system", "content": _INSIGHTS_PROMPT},
                {"role": "user", "content": _stats_block(data["stats"])},
            ],
            max_tokens=900,
            temperature=0.3,
        )
        text = (text or "").strip()
        if text:
            _cache_put(_ai_cache, key, text)
            await _db_cache_put(session, KIND_AI_INSIGHTS, text=text)
        return text
    except Exception as e:
        logger.warning(f"missions_report: AI insights failed, continuing without: {e}")
        return ""


# ── Workbook ───────────────────────────────────────────────────────────────

def build_workbook(data: dict, ai_text: str = "") -> bytes:
    """Build the 3-sheet XLSX. Blocking (openpyxl) — call via run_in_executor."""
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    stats = data["stats"]
    meta = data["meta"]

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    title_font = Font(bold=True, size=15, color="1F4E79")
    section_font = Font(bold=True, size=12, color="1F4E79")
    overdue_fill = PatternFill("solid", fgColor="FCE4E4")
    at_risk_fill = PatternFill("solid", fgColor="FFF3CD")
    wrap = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()

    # ── Sheet 1: summary + insights ────────────────────────────────────
    ws = wb.active
    ws.title = SHEET_SUMMARY
    ws.sheet_view.rightToLeft = True
    for col, width in zip("ABCDEFG", (34, 16, 16, 16, 16, 16, 16)):
        ws.column_dimensions[col].width = width

    r = 1
    ws.cell(r, 1, "חדר מבצעים — דוח מצב").font = title_font
    r += 1
    ws.cell(r, 1, f"הופק בתאריך {meta['generated_at']}").font = Font(italic=True, size=10)
    r += 2

    ws.cell(r, 1, "מדדים מרכזיים").font = section_font
    r += 1
    kpis = [
        ("משימות פתוחות", stats["total_open"]),
        ("באיחור", stats["overdue"]),
        ("אחוז איחור", f"{stats['overdue_rate']}%"),
        ("גיל ממוצע של משימה פתוחה (ימים)", stats["avg_age_open"]),
        ("זמן ביצוע ממוצע, 30 יום (ימים)", stats["avg_cycle_30d"]),
        ("נסגרו ב-7 הימים האחרונים", stats["closed_7d"]),
        ("נפתחו ב-7 הימים האחרונים", stats["created_7d"]),
        (f"משימות ללא עדכון מעל {STALE_DAYS} ימים", stats["stale"]),
        (f"אחראים ללא דיווח מעל {STALE_DAYS} ימים", stats["silent_owners"]),
    ]
    for label, value in kpis:
        ws.cell(r, 1, label).font = Font(bold=True)
        ws.cell(r, 2, value)
        r += 1
    r += 1

    ws.cell(r, 1, "פילוח לפי תאריך יעד").font = section_font
    r += 1
    for b in BUCKET_ORDER:
        ws.cell(r, 1, b)
        ws.cell(r, 2, stats["buckets"].get(b, 0))
        r += 1
    r += 1

    ws.cell(r, 1, "פילוח לפי רביע").font = section_font
    r += 1
    for key, emoji, verb, axis in oms.QUADRANTS:
        ws.cell(r, 1, f"{emoji} {verb} ({axis})")
        ws.cell(r, 2, stats["quadrants"].get(key, 0))
        r += 1
    r += 1

    ws.cell(r, 1, "עומס לפי אחראי").font = section_font
    r += 1
    owner_hdr_row = r
    for c, label in enumerate(
        ["אחראי", "פתוחות", "באיחור", "% איחור", "הושלמו 30 יום", "ממוצע ימי ביצוע",
         "עדכון אחרון"], 1
    ):
        cell = ws.cell(r, c, label)
        cell.font = hdr_font
        cell.fill = hdr_fill
    r += 1
    owner_first_row = r
    for row in stats["owners"]:
        ws.cell(r, 1, row["owner"])
        ws.cell(r, 2, row["open"])
        ws.cell(r, 3, row["overdue"])
        ws.cell(r, 4, row["overdue_rate"])
        ws.cell(r, 5, row["done_30"])
        ws.cell(r, 6, row["avg_cycle"])
        # When the owner last reported on their open work — "—" means never.
        ws.cell(r, 7, row["last_update"])
        if row["overdue"]:
            for c in range(1, 8):
                ws.cell(r, c).fill = overdue_fill
        r += 1
    owner_last_row = r - 1

    if stats["owners"]:
        chart = BarChart()
        chart.type = "col"
        chart.title = "עומס לפי אחראי — פתוחות מול באיחור"
        chart.height, chart.width = 8, 18
        chart.add_data(
            Reference(ws, min_col=2, max_col=3, min_row=owner_hdr_row, max_row=owner_last_row),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(ws, min_col=1, min_row=owner_first_row, max_row=owner_last_row)
        )
        ws.add_chart(chart, f"I{owner_hdr_row}")
    r += 1

    ws.cell(r, 1, "תובנות").font = section_font
    r += 1
    for ins in compute_insights(stats):
        ws.cell(r, 1, f"{ins['icon']} {ins['headline']}").font = Font(bold=True)
        ws.cell(r, 2, ins["action"]).alignment = wrap
        r += 1
    r += 1

    ws.cell(r, 1, "תובנות AI").font = section_font
    r += 1
    if ai_text:
        for line in ai_text.splitlines():
            if line.strip():
                ws.cell(r, 1, line.strip()).alignment = wrap
                r += 1
    else:
        ws.cell(r, 1, "תובנות ה-AI אינן זמינות כרגע — התובנות המחושבות שלמעלה מלאות.")
        r += 1

    # ── Sheet 2: open missions ─────────────────────────────────────────
    ws2 = wb.create_sheet(SHEET_OPEN)
    ws2.sheet_view.rightToLeft = True
    for i, (label, width) in enumerate(zip(OPEN_HEADERS, OPEN_WIDTHS), 1):
        cell = ws2.cell(1, i, label)
        cell.font = hdr_font
        cell.fill = hdr_fill
        ws2.column_dimensions[get_column_letter(i)].width = width
    for ri, row in enumerate(data["open_rows"], start=2):
        values = [
            row["id"], row["title"], row["description"], row["quadrant"], row["urgent"],
            row["important"], row["status"], row["owner"], row["created_by"], row["due"],
            row["days_to_due"], row["days_overdue"], row["age_days"],
            row["created_at"], row["updated_at"], row["last_status_update"],
        ]
        for ci, value in enumerate(values, 1):
            ws2.cell(ri, ci, value)
        fill = overdue_fill if row["_overdue"] else (at_risk_fill if row["_at_risk"] else None)
        if fill:
            for ci in range(1, len(OPEN_HEADERS) + 1):
                ws2.cell(ri, ci).fill = fill
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(OPEN_HEADERS))}{max(1, len(data['open_rows']) + 1)}"

    # ── Sheet 3: closed missions ───────────────────────────────────────
    ws3 = wb.create_sheet(SHEET_CLOSED)
    ws3.sheet_view.rightToLeft = True
    for i, (label, width) in enumerate(zip(CLOSED_HEADERS, CLOSED_WIDTHS), 1):
        cell = ws3.cell(1, i, label)
        cell.font = hdr_font
        cell.fill = hdr_fill
        ws3.column_dimensions[get_column_letter(i)].width = width
    for ri, row in enumerate(data["closed_rows"], start=2):
        values = [
            row["id"], row["title"], row["quadrant"], row["status"], row["owner"],
            row["due"], row["completed_at"], row["cycle_days"], row["on_time"],
            row["created_at"],
        ]
        for ci, value in enumerate(values, 1):
            ws3.cell(ri, ci, value)
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(CLOSED_HEADERS))}{max(1, len(data['closed_rows']) + 1)}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def build_report_bytes(
    session: AsyncSession,
    with_ai: bool = True,
    force: bool = False,
    refresh_data: bool = False,
) -> tuple[bytes, str, str]:
    """Collect → insights → workbook. Returns (xlsx_bytes, filename, generated_at).

    Three modes:
      default          — serve the day's cached report (a dict lookup).
      force            — the 04:10 prewarm: rebuild everything, including the LLM call.
      refresh_data     — rebuild the board data but reuse the day's cached AI narrative,
                         so a manual refresh costs zero Groq tokens.

    `with_ai=False` is an explicit no-LLM download and never touches the cache.
    """
    import asyncio
    from time import perf_counter

    key = oms.today_il().isoformat()
    if with_ai and not force and not refresh_data:
        if key in _report_cache:
            logger.info("⏱ missions report: hit (memory)")
            return _report_cache[key]
        row = await _db_cache_get(session, KIND_REPORT)
        if row is not None and row.blob:
            logger.info("⏱ missions report: hit (db)")
            return _cache_put(
                _report_cache, key,
                (row.blob, row.filename or "", row.generated_at or ""),
            )

    t0 = perf_counter()
    data = await collect_report_data(session)

    if not with_ai:
        ai_text = ""
    elif refresh_data:
        # Reuse today's narrative; only build one if the prewarm never ran.
        ai_text = _ai_cache.get(key) or await get_ai_insights(data, session=session)
    else:
        ai_text = await get_ai_insights(data, force=force, session=session)

    payload = await asyncio.get_event_loop().run_in_executor(
        None, build_workbook, data, ai_text
    )
    result = (
        payload,
        f"דוח_חדר_מבצעים_{data['meta']['date_slug']}.xlsx",
        data["meta"]["generated_at"],
    )
    logger.info(f"⏱ missions report: miss, built in {int((perf_counter() - t0) * 1000)}ms")
    if with_ai:
        _cache_put(_report_cache, key, result)
        if not refresh_data:
            await _db_cache_put(
                session, KIND_REPORT,
                blob=result[0], filename=result[1], generated_at=result[2],
            )
    return result


async def prewarm_daily(session: AsyncSession) -> dict:
    """Build both artifacts and fill the day's caches. Called by the 04:10 cron.

    Each half is guarded separately so a failure in one still leaves the other warm.
    A True here means the cache genuinely holds an LLM-built artifact — the
    summary leg runs strict so a silent fallback is reported as a failure and the
    watchdog picks it up later, instead of the board staying cold all day.
    """
    out = {"report": False, "summary": False}
    try:
        await build_report_bytes(session, with_ai=True, force=True)
        out["report"] = True
    except Exception as e:
        logger.exception(f"missions_report: prewarm of the XLSX report failed: {e}")
    try:
        await build_focus_summary(session, force=True, strict=True)
        out["summary"] = True
    except SummaryNotBuilt:
        logger.warning(
            "missions_report: prewarm of the focus summary fell back to the plain "
            "listing (LLM unavailable) — leaving it cold for the watchdog to retry"
        )
    except Exception as e:
        logger.exception(f"missions_report: prewarm of the focus summary failed: {e}")
    return out


# ── Focus summary (the סיכום AI button) ────────────────────────────────────

_SUMMARY_PROMPT = f"""אתה ראש חדר מבצעים באגף תשתיות חשמל.
לפניך רשימת משימות בסיכון: באיחור, ליעד היום, או ליעד בתוך {AT_RISK_DAYS} ימים.

החזר בעברית בלבד, בדיוק בארבעת הקטעים הבאים, לפי הסדר, עם הכותרות כלשונן:

תמונת מצב
• 3 עד 5 נקודות. כל נקודה מצביעה על דפוס — ריכוז עומס אצל אדם מסוים, אשכול
  משימות באותה תחנה או נושא, איחור שהולך ומעמיק, או צוואר בקבוק חוזר.
• ציין מספרים ושמות מתוך הנתונים. אל תחזור על אותה עובדה פעמיים.

משימות לביצוע היום
• 4 עד 8 נקודות, מסודרות מהדחוף לפחות דחוף.
• כל נקודה בפורמט: [שם האחראי] פועל בהווה + מה בדיוק לעשות — ואז נימוק קצר
  אחרי מקף (כמה ימים באיחור, או מה חוסם).
• התחל כל נקודה בפועל פעולה: תאם, בדוק, סגור, עדכן, הזמן, אשר, הסלם.
• אחד את משימות שאפשר לבצע יחד לנקודה אחת, וציין זאת.

חסמים ותלויות
• 1 עד 3 נקודות: מה עלול למנוע סגירה היום. אם אין — כתוב: אין חסמים ידועים.

לטיפול הנהלה
• רק משימות באיחור של יותר משבוע, או דחופות וחשובות שלא זזות.
• לכל נקודה כתוב מה ההחלטה הנדרשת מההנהלה. אם אין — כתוב: אין.

כללים מחייבים:
- כל שורה ברשימה מתחילה בתו • ורווח. אין פסקאות רצות, אין מספור.
- כל נקודה היא שורה אחת, עד 25 מילים.
- התבסס אך ורק על המשימות שניתנו. אל תמציא משימות, שמות, תאריכים או מספרים.
- אל תשתמש במרכאות כפולות (") — השתמש בגרש בודד (׳) בלבד.
- ללא markdown, ללא כוכביות, ללא תגיות HTML.
"""


async def collect_at_risk(session: AsyncSession) -> dict[str, list[Mission]]:
    """Active missions due within AT_RISK_DAYS (overdue included), bucketed by urgency."""
    today = oms.today_il()
    horizon = today + datetime.timedelta(days=AT_RISK_DAYS)

    missions = list((await session.scalars(
        select(Mission)
        .options(selectinload(Mission.owner), selectinload(Mission.updates))
        .where(
            Mission.status.in_(oms.ACTIVE_STATUSES),
            or_(
                Mission.due_date <= horizon,
                # Undated fires get surfaced too — they have no deadline to trip on.
                (Mission.due_date.is_(None)) & Mission.is_urgent.is_(True)
                & Mission.is_important.is_(True),
            ),
        )
    )).all())

    out: dict[str, list[Mission]] = {"late": [], "today": [], "soon": [], "nodate": []}
    for m in missions:
        if m.due_date is None:
            out["nodate"].append(m)
            continue
        delta = (m.due_date - today).days
        if delta < 0:
            out["late"].append(m)
        elif delta == 0:
            out["today"].append(m)
        else:
            out["soon"].append(m)
    for group in out.values():
        group.sort(key=lambda m: (_priority_weight(m), m.due_date or datetime.date.max))
    return out


_BUCKET_TITLES = {
    "late": "⚠️ באיחור",
    "today": "📌 ליעד היום",
    "soon": f"🟡 בתוך {AT_RISK_DAYS} ימים",
    "nodate": "🔥 דחוף וחשוב, ללא תאריך יעד",
}


def _format_at_risk_plain(groups: dict[str, list[Mission]], today: datetime.date) -> str:
    """Deterministic fallback — used whenever the LLM is unavailable.

    Mirrors the AI layout (bullets + an explicit to-do list) so a Groq outage
    degrades the wording, not the structure.
    """
    if not any(groups.values()):
        return ("‏🟢 <b>אין משימות באיחור או בסיכון</b>\n"
                f"כל המשימות הפעילות עם יעד רחוק מ-{AT_RISK_DAYS} ימים.")

    lines = ["‏🧠 <b>סיכום משימות בסיכון</b>", ""]

    lines.append("<b>🔎 תמונת מצב</b>")
    load: dict[str, int] = {}
    for group in groups.values():
        for m in group:
            load[_user_name(m.owner)] = load.get(_user_name(m.owner), 0) + 1
    total = sum(load.values())
    lines.append(f"• {total} משימות בסיכון: {len(groups['late'])} באיחור, "
                 f"{len(groups['today'])} ליעד היום, {len(groups['soon'])} בתוך {AT_RISK_DAYS} ימים.")
    for name, n in sorted(load.items(), key=lambda kv: -kv[1])[:3]:
        lines.append(f"• {_html.escape(name)} מחזיק {n} מתוך {total} המשימות בסיכון.")
    deep = [m for m in groups["late"] if m.due_date and (today - m.due_date).days > 7]
    if deep:
        lines.append(f"• {len(deep)} משימות באיחור של מעל שבוע.")

    lines.append("")
    lines.append("<b>✅ משימות לביצוע היום</b>")
    ordered = groups["late"] + groups["today"] + groups["soon"] + groups["nodate"]
    for m in ordered[:8]:
        owner = _html.escape(_user_name(m.owner))
        title = _html.escape(m.title or "")
        if m.due_date and m.due_date < today:
            why = f"באיחור של {(today - m.due_date).days} ימים"
        elif m.due_date == today:
            why = "היעד היום"
        elif m.due_date:
            why = f"יעד {oms.format_due(m.due_date)}"
        else:
            why = "דחוף וחשוב, ללא יעד"
        lines.append(f"• [{owner}] קדם את «{title}» — {why}.")

    if deep:
        lines.append("")
        lines.append("<b>🚨 לטיפול הנהלה</b>")
        for m in deep[:5]:
            lines.append(
                f"• {_html.escape(m.title or '')} — באיחור של "
                f"{(today - m.due_date).days} ימים אצל {_html.escape(_user_name(m.owner))}."
            )

    lines.append("")
    lines.append("<i>סיכום ה-AI אינו זמין כרגע — זו הרשימה המחושבת.</i>")
    return "\n".join(lines)


def _at_risk_context(groups: dict[str, list[Mission]], today: datetime.date) -> str:
    """Everything the model needs to spot patterns: owner, age, depth of delay, context."""
    lines = []

    # Owner load first — this is what lets the model say "X holds most of the delay".
    load: dict[str, int] = {}
    for group in groups.values():
        for m in group:
            load[_user_name(m.owner)] = load.get(_user_name(m.owner), 0) + 1
    if load:
        lines.append("עומס בסיכון לפי אחראי:")
        for name, n in sorted(load.items(), key=lambda kv: -kv[1]):
            lines.append(f"- {_sanitize(name)}: {n} משימות")
        lines.append("")

    # The prompt used to grow linearly with the board — ~300 chars of detail per
    # mission with no cap, which is what made a cold build slow on a busy board.
    # Buckets are walked worst-first, so the cap only ever drops the least urgent.
    budget = PROMPT_MISSION_CAP
    for key in ("late", "today", "soon", "nodate"):
        group = (groups.get(key) or [])[:budget]
        if not group:
            continue
        budget -= len(group)
        lines.append(f"{_BUCKET_TITLES[key].split(' ', 1)[-1]}:")
        for m in group:
            due = oms.format_due(m.due_date) if m.due_date else "ללא יעד"
            bits = [
                f"אחראי: {_sanitize(_user_name(m.owner))}",
                f"יעד: {due}",
                f"סטטוס: {oms.STATUS_LABELS.get(m.status, m.status)}",
                oms.quadrant_label(oms.quadrant_key(m), with_axis=True),
            ]
            if m.due_date and m.due_date < today:
                bits.append(f"באיחור של {(today - m.due_date).days} ימים")
            age = _age_days(m.created_at, today)
            if age is not None:
                bits.append(f"נפתחה לפני {age} ימים")
            line = f"- {_sanitize(m.title)} | " + " | ".join(bits)
            detail = _description_with_updates(m)
            if detail:
                line += f" | פירוט: {_sanitize(detail)[:300]}"
            lines.append(line)
        lines.append("")
    return "\n".join(lines)


_SECTION_ICONS = {
    "תמונת מצב": "🔎",
    "משימות לביצוע היום": "✅",
    "חסמים ותלויות": "⛔",
    "לטיפול הנהלה": "🚨",
}


def _decorate_summary(raw: str) -> str:
    """Bold the model's section headers and normalise its bullets, after escaping."""
    out = []
    for line in _html.escape(raw).splitlines():
        stripped = line.strip()
        if not stripped:
            out.append("")
            continue
        if stripped in _SECTION_ICONS:
            out.append("")
            out.append(f"<b>{_SECTION_ICONS[stripped]} {stripped}</b>")
            continue
        # The model occasionally emits "-" or "*" despite the instruction.
        if stripped[0] in "-*·" and len(stripped) > 1:
            stripped = "• " + stripped[1:].strip()
        out.append(stripped)
    return "\n".join(out).strip()


class SummaryNotBuilt(Exception):
    """The LLM leg of the focus summary failed and the plain fallback was served.

    Raised only towards prewarm_daily, so a prewarm can no longer report success
    while leaving the cache empty — the exact bug that made the 🧠 סיכום AI button
    slow all day after a bad 04:10.
    """

    def __init__(self, fallback: str):
        super().__init__("focus summary fell back to the plain listing")
        self.fallback = fallback


async def build_focus_summary(
    session: AsyncSession, force: bool = False, strict: bool = False,
) -> str:
    """AI situation assessment + to-do list for everything overdue or due within AT_RISK_DAYS.

    Served from the day cache (memory, then the DB mirror); the 04:10 prewarm and
    the watchdog build it with force=True. Always returns sendable HTML: falls back
    to the computed listing if the LLM fails. `strict=True` additionally raises
    SummaryNotBuilt on that fallback, so callers that care whether the cache is now
    warm can tell — plain callers still just get the fallback text.
    """
    from time import perf_counter

    key = oms.today_il().isoformat()
    if not force:
        if key in _summary_cache:
            logger.info("⏱ missions summary: hit (memory)")
            return _summary_cache[key]
        row = await _db_cache_get(session, KIND_SUMMARY)
        if row is not None and row.text:
            logger.info("⏱ missions summary: hit (db)")
            return _cache_put(_summary_cache, key, row.text)

    t0 = perf_counter()
    today = oms.today_il()
    groups = await collect_at_risk(session)
    if not any(groups.values()):
        plain = _format_at_risk_plain(groups, today)
        _cache_put(_summary_cache, key, plain)
        await _db_cache_put(session, KIND_SUMMARY, text=plain)
        return plain

    def _fallback() -> str:
        # Not cached — a transient outage shouldn't pin the fallback for the whole day.
        text = _format_at_risk_plain(groups, today)
        if strict:
            raise SummaryNotBuilt(text)
        return text

    from app.services.llm_router import llm_chat
    try:
        raw = await llm_chat(
            "missions_summary",
            [
                {"role": "system", "content": _SUMMARY_PROMPT},
                {"role": "user", "content": _at_risk_context(groups, today)},
            ],
            max_tokens=1600,
            temperature=0.3,
        )
    except Exception as e:
        logger.warning(f"missions_report: focus summary LLM failed, using plain list: {e}")
        return _fallback()

    raw = (raw or "").strip()
    if not raw:
        logger.warning("missions_report: focus summary LLM returned empty, using plain list")
        return _fallback()

    counts = (
        f"⚠️ {len(groups['late'])} באיחור · "
        f"📌 {len(groups['today'])} להיום · "
        f"🟡 {len(groups['soon'])} בתוך {AT_RISK_DAYS} ימים"
    )
    stamp = datetime.datetime.now(oms._IL_TZ).strftime("%d/%m/%Y %H:%M")
    # Escape the model output before wrapping it in our own tags — it must not inject markup.
    text = (f"‏🧠 <b>סיכום משימות בסיכון</b>\n<i>{counts}</i>\n"
            f"<i>נכון ל-{stamp}</i>\n\n{_decorate_summary(raw)}")

    from app.services.telegram_routing import _TG_MAX
    if len(text) > _TG_MAX:
        text = text[: _TG_MAX - 20] + "\n…(קוצר)"
    logger.info(f"⏱ missions summary: miss, built in {int((perf_counter() - t0) * 1000)}ms")
    _cache_put(_summary_cache, key, text)
    await _db_cache_put(session, KIND_SUMMARY, text=text)
    return text
